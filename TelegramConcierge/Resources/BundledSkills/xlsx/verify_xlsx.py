#!/usr/bin/env python3
"""Verify .xlsx workbooks: structure, content, diffs, formula and data lint.

Subcommands:
  manifest book.xlsx
      Workbook-level structure: sheets, dimensions, tables, charts, images,
      defined names, hidden sheets. Run before AND after editing an existing
      workbook — any difference must be intentional and explainable.

  inspect book.xlsx [--sheet NAME] [--rows N]
      Per-sheet headers plus the first N data rows as
      (value, data_type, number_format) triples. Verifies types and formats,
      not just values.

  diff before.xlsx after.xlsx [--sheet NAME] [--limit N]
      Cell-by-cell value diff (formulas as text, not cached results).
      Confirms edits are limited to the intended cells.

  check book.xlsx [--no-eval]
      Formula and data-integrity lint:
      - formulas referencing nonexistent sheets, empty areas, or containing #REF!
      - aggregation ranges that stop short of adjacent data (SUM(B2:B9)
        with data continuing in B10)
      - formulas accidentally stored as text
      - numbers/dates stored as text, header whitespace, merged cells and
        blank rows inside data regions
      - when the 'formulas' package is installed: full workbook evaluation,
        reporting #REF!/#DIV/0!/#VALUE!/#NAME? results and computed values
        of formula cells so totals can be sanity-checked

All output is JSON. Findings are leads for review, not verdicts — judge
each against the workbook's intent.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("error: openpyxl is required (python3 -m pip install openpyxl)", file=sys.stderr)
    raise SystemExit(1)


def build_manifest(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    try:
        defined_names = sorted(wb.defined_names.keys())
    except AttributeError:  # older openpyxl API
        defined_names = sorted(dn.name for dn in wb.defined_names.definedName)
    return {
        "sheets": wb.sheetnames,
        "defined_names": defined_names,
        "tables": {ws.title: sorted(ws.tables.keys()) for ws in wb.worksheets},
        "charts": {ws.title: len(ws._charts) for ws in wb.worksheets},
        "images": {ws.title: len(ws._images) for ws in wb.worksheets},
        "dimensions": {ws.title: [ws.max_row, ws.max_column] for ws in wb.worksheets},
        "hidden_sheets": [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"],
        "frozen_panes": {ws.title: ws.freeze_panes for ws in wb.worksheets if ws.freeze_panes},
    }


def inspect(path: Path, sheet: str | None, rows: int) -> dict:
    wb = load_workbook(path, data_only=False)
    sheets = [sheet] if sheet else wb.sheetnames
    out: dict = {}
    for name in sheets:
        if name not in wb.sheetnames:
            out[name] = {"error": "sheet not found"}
            continue
        ws = wb[name]
        headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        sample = []
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 1 + rows)):
            sample.append([
                {"value": c.value, "type": c.data_type, "format": c.number_format}
                for c in row
            ])
        out[name] = {
            "rows": ws.max_row,
            "cols": ws.max_column,
            "headers": headers,
            "sample_rows": sample,
        }
    return out


def diff(before_path: Path, after_path: Path, sheet: str | None, limit: int) -> dict:
    before = load_workbook(before_path, data_only=False)
    after = load_workbook(after_path, data_only=False)
    changes: list = []
    result: dict = {
        "sheets_removed": [s for s in before.sheetnames if s not in after.sheetnames],
        "sheets_added": [s for s in after.sheetnames if s not in before.sheetnames],
        "cell_changes": changes,
    }
    sheets = [sheet] if sheet else [s for s in before.sheetnames if s in after.sheetnames]
    truncated = False
    for name in sheets:
        if name not in before.sheetnames or name not in after.sheetnames:
            continue
        ws0, ws1 = before[name], after[name]
        for row in range(1, max(ws0.max_row, ws1.max_row) + 1):
            for col in range(1, max(ws0.max_column, ws1.max_column) + 1):
                a = ws0.cell(row, col).value
                b = ws1.cell(row, col).value
                if a != b:
                    if len(changes) >= limit:
                        truncated = True
                        break
                    changes.append({
                        "sheet": name,
                        "cell": ws1.cell(row, col).coordinate,
                        "before": a,
                        "after": b,
                    })
            if truncated:
                break
        if truncated:
            break
    result["total_shown"] = len(changes)
    result["truncated"] = truncated
    return result


# --- check: formula + data lint ---------------------------------------------

import re

# Sheet-qualified reference: 'My Sheet'!A1:B10 or Sheet1!A1
SHEET_REF_RE = re.compile(
    r"(?:'([^']+)'|(\b[A-Za-z_][\w.]*))!"
    r"(\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?|\$?[A-Z]{1,3}:\$?[A-Z]{1,3})"
)
# Bare reference: A1 or A1:B10 (not preceded by sheet!/word, not a function call)
BARE_REF_RE = re.compile(
    r"(?<![!\w:$])(\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)(?![\w(:])"
)
NUMERIC_TEXT_RE = re.compile(r"^-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?$")
DATE_TEXT_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$|^\d{1,2}[/.]\d{1,2}[/.]\d{2,4}$")


def strip_formula_strings(formula: str) -> str:
    return re.sub(r'"[^"]*"', '""', formula)


def formula_references(formula: str, own_sheet: str) -> list:
    """Extract (sheet, range) pairs from a formula, resolving bare refs to own_sheet."""
    body = strip_formula_strings(formula)
    refs = []
    for match in SHEET_REF_RE.finditer(body):
        sheet = match.group(1) or match.group(2)
        refs.append((sheet, match.group(3)))
    without_qualified = SHEET_REF_RE.sub("", body)
    for match in BARE_REF_RE.finditer(without_qualified):
        refs.append((own_sheet, match.group(1)))
    return refs


def check_workbook(path: Path, evaluate: bool) -> dict:
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(path, data_only=False)
    sheet_names = {name.lower(): name for name in wb.sheetnames}
    errors: list = []
    warnings: list = []
    formula_cells: list = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=") and cell.data_type == "f"):
                    formula_cells.append((ws, cell))
                elif isinstance(cell.value, str) and cell.value.lstrip().startswith("=") and cell.data_type == "s":
                    errors.append({
                        "kind": "formula_stored_as_text",
                        "sheet": ws.title, "cell": cell.coordinate,
                        "detail": f"Cell contains formula text {cell.value[:60]!r} but is stored as a string — it will never calculate.",
                    })

    for ws, cell in formula_cells:
        formula = str(cell.value)
        if "#REF!" in formula:
            errors.append({
                "kind": "broken_reference",
                "sheet": ws.title, "cell": cell.coordinate,
                "detail": f"Formula contains #REF!: {formula[:80]}",
            })
        for ref_sheet, ref_range in formula_references(formula, ws.title):
            target_name = sheet_names.get(ref_sheet.lower())
            if target_name is None:
                errors.append({
                    "kind": "missing_sheet",
                    "sheet": ws.title, "cell": cell.coordinate,
                    "detail": f"Formula references sheet '{ref_sheet}' which does not exist: {formula[:80]}",
                })
                continue
            target = wb[target_name]
            try:
                min_col, min_row, max_col, max_row = range_boundaries(ref_range.replace("$", ""))
            except Exception:
                continue
            if min_row is None or max_row is None:
                continue  # whole-column ref — always fine
            if min_row > target.max_row or min_col > target.max_column:
                warnings.append({
                    "kind": "empty_area_reference",
                    "sheet": ws.title, "cell": cell.coordinate,
                    "detail": f"{ref_range} on '{target_name}' lies entirely outside the used area "
                              f"({target.max_row}x{target.max_column}).",
                })
                continue
            # Single-column vertical range: does data continue right after it?
            if min_col == max_col and max_row > min_row:
                next_cell = target.cell(row=max_row + 1, column=min_col)
                is_own = target_name == ws.title and next_cell.coordinate == cell.coordinate
                if not is_own and next_cell.value is not None and next_cell.data_type != "f":
                    warnings.append({
                        "kind": "range_stops_short",
                        "sheet": ws.title, "cell": cell.coordinate,
                        "detail": f"{ref_range} ends at row {max_row} but '{target_name}'!{next_cell.coordinate} "
                                  f"holds more data ({str(next_cell.value)[:30]!r}). Did the range miss appended rows?",
                    })

    # Data-integrity lint
    numbers_as_text: list = []
    dates_as_text: list = []
    header_whitespace: list = []
    blank_rows: dict = {}
    merged: dict = {}
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                if isinstance(cell.value, str) and cell.value != cell.value.strip():
                    header_whitespace.append({"sheet": ws.title, "cell": cell.coordinate, "value": cell.value})
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.data_type == "s" and isinstance(cell.value, str):
                    text = cell.value.strip()
                    if NUMERIC_TEXT_RE.match(text) and not (len(text) > 1 and text.startswith("0") and "." not in text):
                        numbers_as_text.append({"sheet": ws.title, "cell": cell.coordinate, "value": text})
                    elif DATE_TEXT_RE.match(text):
                        dates_as_text.append({"sheet": ws.title, "cell": cell.coordinate, "value": text})
        empty = [r[0].row for r in ws.iter_rows(min_row=2, max_row=ws.max_row)
                 if all(c.value is None for c in r)] if ws.max_row >= 2 else []
        # Only blank rows with data after them are suspicious.
        empty = [r for r in empty if r < ws.max_row]
        if empty:
            blank_rows[ws.title] = empty[:20]
        ranges = [str(r) for r in ws.merged_cells.ranges]
        if ranges:
            merged[ws.title] = ranges[:20]

    report: dict = {
        "sheets": wb.sheetnames,
        "formulas_found": len(formula_cells),
        "errors": errors,
        "warnings": warnings,
        "data_lint": {
            "numbers_stored_as_text": numbers_as_text[:25],
            "dates_stored_as_text": dates_as_text[:25],
            "header_whitespace": header_whitespace,
            "blank_rows_inside_data": blank_rows,
            "merged_cell_ranges": merged,
            "note": "Leading-zero identifiers are correctly text and are not flagged. "
                    "Merged ranges in title rows are normal; merges inside data tables break sorting/filtering.",
        },
    }

    if evaluate and formula_cells:
        targets = {(ws.title.upper(), cell.coordinate) for ws, cell in formula_cells}
        report["evaluation"] = evaluate_workbook(path, targets)
    elif formula_cells:
        report["evaluation"] = {"skipped": "--no-eval"}
    else:
        report["evaluation"] = {"skipped": "no formulas in workbook"}

    return report


def evaluate_workbook(path: Path, formula_targets: set) -> dict:
    try:
        import formulas  # type: ignore
    except ImportError:
        return {
            "skipped": "the 'formulas' package is not installed",
            "enable": "python3 -m pip install formulas",
        }

    import contextlib
    import io
    import logging

    try:
        logging.disable(logging.WARNING)
        with contextlib.redirect_stderr(io.StringIO()):  # silence tqdm progress bars
            model = formulas.ExcelModel().loads(str(path)).finish()
            solution = model.calculate()
    except Exception as error:
        return {"failed": f"evaluation error: {error}"}
    finally:
        logging.disable(logging.NOTSET)

    computed: list = []
    eval_errors: list = []
    for key, value in solution.items():
        # Keys look like "'[BOOK.XLSX]SHEET'!A1". Skip non-cell entries and
        # plain data cells — only formula results are interesting.
        match = re.match(r"^'\[[^\]]+\]([^']+)'!([A-Z]+\d+)$", key)
        if not match or (match.group(1).upper(), match.group(2)) not in formula_targets:
            continue
        try:
            scalar = value.value[0, 0] if hasattr(value, "value") else value
        except Exception:
            scalar = str(value)
        text = str(scalar)
        entry = {"sheet": match.group(1), "cell": match.group(2), "value": text[:80]}
        if text.startswith("#"):
            eval_errors.append(entry)
        else:
            computed.append(entry)

    return {
        "engine": "formulas",
        "error_cells": eval_errors,
        "computed_formulas": computed[:50],
        "computed_total": len(computed),
        "note": "error_cells must be empty. Sanity-check computed totals against expectations.",
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Verify .xlsx workbooks.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_manifest = sub.add_parser("manifest")
    p_manifest.add_argument("book", type=Path)

    p_inspect = sub.add_parser("inspect")
    p_inspect.add_argument("book", type=Path)
    p_inspect.add_argument("--sheet")
    p_inspect.add_argument("--rows", type=int, default=3)

    p_diff = sub.add_parser("diff")
    p_diff.add_argument("before", type=Path)
    p_diff.add_argument("after", type=Path)
    p_diff.add_argument("--sheet")
    p_diff.add_argument("--limit", type=int, default=200)

    p_check = sub.add_parser("check")
    p_check.add_argument("book", type=Path)
    p_check.add_argument("--no-eval", action="store_true", help="Skip full formula evaluation")

    args = parser.parse_args()

    paths = [getattr(args, k) for k in ("book", "before", "after") if hasattr(args, k)]
    for p in paths:
        if not p.expanduser().exists():
            print(f"error: file not found: {p}", file=sys.stderr)
            return 2

    if args.command == "manifest":
        out = build_manifest(args.book.expanduser())
    elif args.command == "inspect":
        out = inspect(args.book.expanduser(), args.sheet, args.rows)
    elif args.command == "check":
        out = check_workbook(args.book.expanduser(), evaluate=not args.no_eval)
    else:
        out = diff(args.before.expanduser(), args.after.expanduser(), args.sheet, args.limit)

    print(json.dumps(out, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
